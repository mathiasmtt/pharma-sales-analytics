"""Tests sobre la generación del reporte Excel."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from src.reporting import build_excel_report, report_filename


class TestBuildExcelReport:
    def test_retorna_bytes(self, daily_sintetico, monthly_sintetico):
        data = build_excel_report(daily_sintetico, monthly_sintetico)
        assert isinstance(data, bytes)
        assert len(data) > 1000

    def test_hojas_esperadas(self, daily_sintetico, monthly_sintetico):
        data = build_excel_report(daily_sintetico, monthly_sintetico)
        wb = load_workbook(BytesIO(data))
        esperadas = {
            "Resumen",
            "Pareto",
            "Tendencia Mensual",
            "YoY",
            "Top Meses",
            "Desvios",
        }
        assert esperadas.issubset(set(wb.sheetnames))

    def test_funciona_sin_monthly(self, daily_sintetico):
        data = build_excel_report(daily_sintetico, monthly=None)
        wb = load_workbook(BytesIO(data))
        assert "YoY" not in wb.sheetnames
        assert "Resumen" in wb.sheetnames


class TestReportFilename:
    def test_formato(self, daily_sintetico):
        name = report_filename(daily_sintetico)
        assert name.startswith("pharma_sales_report_")
        assert name.endswith(".xlsx")

    def test_incluye_rango(self, daily_sintetico):
        name = report_filename(daily_sintetico)
        ini = daily_sintetico["date"].min().strftime("%Y%m%d")
        fin = daily_sintetico["date"].max().strftime("%Y%m%d")
        assert ini in name
        assert fin in name
